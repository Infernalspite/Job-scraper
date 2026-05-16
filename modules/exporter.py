import openpyxl
from openpyxl.styles import Font, PatternFill

def export_to_excel(jobs, output_path="matched_jobs.xlsx"):
    if not jobs:
        print("No jobs to export")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matched Jobs"

    headers = ["Job/Internship Name", "Company", "Description", "Last Date", "Match Score", "Salary"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    # Add data
    for job in jobs:
        ws.append([
            job.get("title", ""),
            job.get("company", ""),
            job.get("description", "")[:300],
            job.get("last_date", ""),
            round(job.get("match_score", 0), 3),
            job.get("salary", 0)
        ])

    wb.save(output_path)
    print("Excel export complete")